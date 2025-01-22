from openpyxl import load_workbook
import pandas as pd

def extraer_tabla_y_limpiar(hoja, rango):
    """
    Extrae los datos de un rango específico en una hoja de Excel y descombina las celdas combinadas.

    Args:
        hoja: Objeto Worksheet de openpyxl.
        rango: Tupla con (min_row, min_col, max_row, max_col).

    Returns:
        Una lista de listas representando la tabla extraída.
    """
    min_row, min_col, max_row, max_col = rango

    # Hacer una copia de los rangos de celdas combinadas
    rangos_combinados = list(hoja.merged_cells.ranges)
    for rango_combinado in rangos_combinados:
        if (rango_combinado.min_row >= min_row and rango_combinado.max_row <= max_row and
            rango_combinado.min_col >= min_col and rango_combinado.max_col <= max_col):
            
            valor = hoja.cell(rango_combinado.min_row, rango_combinado.min_col).value
            hoja.unmerge_cells(rango_combinado.coord)

            for fila in hoja.iter_rows(min_row=rango_combinado.min_row, max_row=rango_combinado.max_row,
                                       min_col=rango_combinado.min_col, max_col=rango_combinado.max_col):
                for celda in fila:
                    celda.value = valor

    tabla = [
        [hoja.cell(row=fila, column=col).value for col in range(min_col, max_col + 1)]
        for fila in range(min_row, max_row + 1)
    ]
    return tabla

def procesar_archivo(archivo, rango_tabla, hoja_nombre="ESC2"):
    """
    Extrae y procesa la tabla de un archivo Excel.
    Retorna un DataFrame con los valores de la tabla.
    """
    wb = load_workbook(archivo)
    hoja = wb[hoja_nombre]
    tabla = extraer_tabla_y_limpiar(hoja, rango_tabla)
    
    df = pd.DataFrame(tabla)
    
    # Renombrar columnas usando la fila de encabezados
    encabezados = df.iloc[1]
    df.columns = encabezados
    df = df.iloc[2:].reset_index(drop=True)
    
    return df

def consolidar_archivos(archivos, rango_sumatoria, rango_tabla, hoja_nombre="ESC2"):
    """
    Consolida los datos de múltiples archivos Excel y retorna un DataFrame con los datos sumados.

    Args:
        archivos (list): Lista de rutas de los archivos.
        rango_sumatoria (tuple): Tupla (min_row, min_col, max_row, max_col) del rango de valores a sumar.
        hoja_nombre (str): Nombre de la hoja donde se encuentra la tabla.

    Returns:
        pd.DataFrame: DataFrame con los valores consolidados.
    """

    # Extraer las coordenadas del rango de sumatoria
    min_row, min_col, max_row, max_col = rango_sumatoria

    # DataFrame consolidado intermedio
    consolidado = None

    for archivo in archivos:
        # Procesar el archivo y extraer la tabla
        df = procesar_archivo(archivo, rango_tabla=rango_tabla, hoja_nombre=hoja_nombre)

        # Seleccionar rango numérico (filas y columnas relevantes)
        rango_datos = df.iloc[min_row - 2 : max_row - 1 , min_col - 1  : max_col  ].copy()  # Ajuste de coordenadas

        
        rango_datos = rango_datos.apply(pd.to_numeric, errors='coerce').fillna(0)  # Convertir a numérico

        print(f"Archivo: {archivo}")
        print(f"Rango seleccionado del DataFrame:")
        print(rango_datos)
        # Sumar al consolidado
        if consolidado is None:
            consolidado = rango_datos
        else:
            consolidado += rango_datos

    return consolidado

from openpyxl import load_workbook
import pandas as pd

def inyectar_datos_en_plantilla(df_consolidado, hoja_plantilla, rango_sumatoria):
    """
    Inyecta los datos consolidados en una plantilla sin alterar el formato original, 
    excepto la última fila con celdas combinadas.

    Args:
        df_consolidado (pd.DataFrame): DataFrame con los datos consolidados.
        hoja_plantilla (openpyxl.worksheet.worksheet.Worksheet): Hoja de la plantilla base.
        rango_sumatoria (tuple): Tupla (min_row, min_col, max_row, max_col) para el rango donde se inyectarán los datos.
    """
    min_row, min_col, max_row, max_col = rango_sumatoria

    # Asegurarse de que el DataFrame tenga suficientes datos para inyectar
    for i, row in enumerate(df_consolidado.itertuples(index=False)):  # Se usa itertuples para evitar el uso del index
        # Limitar la cantidad de filas a inyectar a las filas disponibles en el rango (max_row) menos la última fila
        if min_row + i >= max_row:  # Evitar inyectar en la última fila
            break

        for j in range(min_col, max_col + 1):  # Ya no es necesario saltar celdas por pares
            hoja_plantilla.cell(row=min_row + i, column=j).value = row[j - min_col]

    # Los valores se inyectan directamente en la plantilla, que ya tiene el formato correcto
