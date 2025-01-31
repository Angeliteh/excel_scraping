from openpyxl import load_workbook
import pandas as pd
import win32com.client

def extraer_tabla_y_limpiar(hoja, rango):
    """
    Extrae los datos de un rango específico en una hoja de Excel y descombina las celdas combinadas.

    Args:
        hoja: Objeto Worksheet de openpyxl.
        rango: Tupla con (min_row, min_col, max_row, max_col).

    Returns:
        Una lista de listas representando la tabla extraída.
    """
    min_row, min_col, max_row, max_col = rango

    # Hacer una copia de los rangos de celdas combinadas
    rangos_combinados = list(hoja.merged_cells.ranges)
    for rango_combinado in rangos_combinados:
        if (rango_combinado.min_row >= min_row and rango_combinado.max_row <= max_row and
            rango_combinado.min_col >= min_col and rango_combinado.max_col <= max_col):
            
            valor = hoja.cell(rango_combinado.min_row, rango_combinado.min_col).value
            hoja.unmerge_cells(rango_combinado.coord)

            for fila in hoja.iter_rows(min_row=rango_combinado.min_row, max_row=rango_combinado.max_row,
                                       min_col=rango_combinado.min_col, max_col=rango_combinado.max_col):
                for celda in fila:
                    celda.value = valor

    tabla = [
        [hoja.cell(row=fila, column=col).value for col in range(min_col, max_col + 1)]
        for fila in range(min_row, max_row + 1)
    ]
    return tabla

def procesar_archivo(archivo, rango_tabla, hoja_nombre="ESC2"):
    """
    Extrae y procesa la tabla de un archivo Excel.
    Retorna un DataFrame con los valores de la tabla.
    """
    wb = load_workbook(archivo)
    hoja = wb[hoja_nombre]
    tabla = extraer_tabla_y_limpiar(hoja, rango_tabla)
    
    df = pd.DataFrame(tabla)
    
    # Renombrar columnas usando la fila de encabezados
    encabezados = df.iloc[1]
    df.columns = encabezados
    df = df.iloc[2:].reset_index(drop=True)
    
    return df

def consolidar_archivos(archivos, rango_sumatoria, rango_tabla, hoja_nombre="ESC2"):
    """
    Consolida los datos de múltiples archivos Excel y retorna un DataFrame con los datos sumados.

    Args:
        archivos (list): Lista de rutas de los archivos.
        rango_sumatoria (tuple): Tupla (min_row, min_col, max_row, max_col) del rango de valores a sumar.
        hoja_nombre (str): Nombre de la hoja donde se encuentra la tabla.

    Returns:
        pd.DataFrame: DataFrame con los valores consolidados.
    """

    # Extraer las coordenadas del rango de sumatoria
    min_row, min_col, max_row, max_col = rango_sumatoria

    # DataFrame consolidado intermedio
    consolidado = None

    for archivo in archivos:
        # Procesar el archivo y extraer la tabla
        df = procesar_archivo(archivo, rango_tabla=rango_tabla, hoja_nombre=hoja_nombre)

        # Seleccionar rango numérico (filas y columnas relevantes)
        rango_datos = df.iloc[min_row - 2 : max_row - 1 , min_col - 1  : max_col  ].copy()  # Ajuste de coordenadas

        
        rango_datos = rango_datos.apply(pd.to_numeric, errors='coerce').fillna(0)  # Convertir a numérico

        print(f"Archivo: {archivo}")
        print(f"Rango seleccionado del DataFrame:")
        print(rango_datos)
        # Sumar al consolidado
        if consolidado is None:
            consolidado = rango_datos
        else:
            consolidado += rango_datos

    return consolidado

def inyectar_datos_en_plantilla(df_consolidado, hoja_plantilla, rango_inyeccion):
    """
    Inyecta los datos consolidados en una plantilla, manejando las celdas combinadas en la última fila.

    Args:
        df_consolidado (pd.DataFrame): DataFrame con los datos consolidados.
        hoja_plantilla (openpyxl.worksheet.worksheet.Worksheet): Hoja de la plantilla base.
        rango_sumatoria (tuple): Tupla (min_row, min_col, max_row, max_col) para el rango donde se inyectarán los datos.
    """
    min_row, min_col, max_row, max_col = rango_inyeccion
    print(f"Inicio de inyección de datos en rango: {min_row}-{max_row}, columnas: {min_col}-{max_col}")  # Depuración

    # Asegurarse de que el DataFrame tenga suficientes datos para inyectar
    for i, row in enumerate(df_consolidado.itertuples(index=False)):  # Se usa itertuples para evitar el uso del index
        print(f"Inyectando datos para la fila {min_row + i}")  # Depuración
        for j in range(min_col, max_col + 1):  # Ya no es necesario saltar celdas por pares
            cell = hoja_plantilla.cell(row=min_row + i, column=j)
            
            # Verificamos si la celda está dentro de un rango de celdas combinadas
            cell_is_merged = False
            for range_ in hoja_plantilla.merged_cells.ranges:
                if (min_row + i >= range_.min_row and min_row + i <= range_.max_row and
                    j >= range_.min_col and j <= range_.max_col):
                    cell_is_merged = True
                    break

            if cell_is_merged:
                print(f"Celda {cell.coordinate} está combinada. Asignando valor solo a la celda superior izquierda.")
                # Encontramos la celda superior izquierda de la celda combinada
                top_left_cell = hoja_plantilla.cell(row=range_.min_row, column=range_.min_col)
                top_left_cell.value = row[j - min_col]
            else:
                print(f"Inyectando valor '{row[j - min_col]}' en celda {cell.coordinate}")  # Depuración
                cell.value = row[j - min_col]

    # Ahora procesamos la última fila (la fila 16) donde las celdas están combinadas
    if min_row + i == max_row - 1:  # Si estamos en la última fila
        print(f"Procesando última fila (fila {max_row})...")  # Depuración
        for col in range(min_col, max_col + 1, 2):  # Iteramos cada dos columnas
            # Descombinamos las celdas combinadas en esa fila
            cell1 = hoja_plantilla.cell(row=max_row, column=col)
            cell2 = hoja_plantilla.cell(row=max_row, column=col + 1)

            # Verificamos si las celdas están combinadas antes de descombinarlas
            if any(cell1.coordinate in range_[0] for range_ in hoja_plantilla.merged_cells.ranges):
                print(f"Descombinando celdas: {cell1.coordinate} y {cell2.coordinate}")
                hoja_plantilla.unmerge_cells(f"{cell1.coordinate}:{cell2.coordinate}")  # Descombinar celdas
            else:
                print(f"Las celdas {cell1.coordinate} y {cell2.coordinate} no están combinadas")

            # Asignamos el valor del DataFrame (para ambos elementos del par de celdas)
            valor = row[col - min_col]  # Tomamos el valor del DataFrame para el par de celdas
            print(f"Asignando valor '{valor}' a las celdas combinadas: {cell1.coordinate} y {cell2.coordinate}")
            cell1.value = valor
            cell2.value = valor

            # Recombinamos las celdas después de asignar los valores
            hoja_plantilla.merge_cells(
                start_row=max_row, start_column=col,
                end_row=max_row, end_column=col + 1
            )
            print(f"Recombinando celdas: {cell1.coordinate} y {cell2.coordinate}")  # Depuración

def inyectar_formulas_totales_y_subtotales(archivo_salida, hoja_nombre="ZONA3", fila_inicial=6, fila_final=13):
    """
    Inyecta fórmulas en un rango de filas para realizar sumas en celdas combinadas
    y calcular totales en una hoja de Excel.

    Args:
        archivo_salida (str): Ruta del archivo Excel donde se inyectarán las fórmulas.
        hoja_nombre (str): Nombre de la hoja donde se aplicarán las fórmulas.
        fila_inicial (int): Fila inicial donde se comenzarán a insertar las fórmulas.
        fila_final (int): Fila final donde se detendrán las inserciones de fórmulas.
    """
    wb = load_workbook(archivo_salida)
    hoja = wb[hoja_nombre]

    for fila in range(fila_inicial, fila_final + 1):
        hoja[f"T{fila}"] = f"=H{fila}+J{fila}+L{fila}+N{fila}+P{fila}+R{fila}"  # Suma para T/U
        hoja[f"V{fila}"] = f"=I{fila}+K{fila}+M{fila}+O{fila}+Q{fila}+S{fila}"  # Suma para V/W
        hoja[f"X{fila}"] = f"=T{fila}+V{fila}"  # Suma de totales en X/Y/Z

    wb.save(archivo_salida)
    print(f"Fórmulas inyectadas correctamente en las filas {fila_inicial} a {fila_final} de: {archivo_salida}")

def convertir_formulas_a_valores(archivo_salida):
    """
    Abre el archivo en Excel, fuerza el cálculo de las fórmulas y guarda solo los valores.
    """
    try:
        # Abrir el archivo con Excel (debe estar instalado en Windows)
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False  # No mostrar Excel
        wb = excel.Workbooks.Open(archivo_salida)

        # Forzar el cálculo de todas las fórmulas
        wb.RefreshAll()
        excel.CalculateUntilAsyncQueriesDone()
        wb.Save()

        # Cerrar Excel
        wb.Close(SaveChanges=True)
        excel.Quit()

        # Ahora, cargar con openpyxl para eliminar las fórmulas
        wb = load_workbook(archivo_salida, data_only=True)
        ws = wb.active  # Usa la hoja activa (ajustar si es otra hoja)

        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell.value = cell.value  # Mantiene solo el valor, elimina la fórmula

        wb.save(archivo_salida)  # Guarda el archivo sin fórmulas
        print(f"Las fórmulas han sido convertidas a valores en: {archivo_salida}")

    except Exception as e:
        print(f"Error al convertir fórmulas a valores: {e}")